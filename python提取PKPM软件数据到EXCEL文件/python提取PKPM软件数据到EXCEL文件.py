"""
python提取PKPM软件数据写入到EXCEL文件.py
http://www.oschina.net/question/2661202_2154877
2016年3月5日 00:34:38 codegay
原文数据很多没说清楚，省点精力就写个大概。
"""
import re
import openpyxl

with open("a.txt",encoding="utf-8") as f:
    txt=f.read()

#表头    
head=["标准层","N-C","工况","Nu","Uc","N-C","MX","MY","N"]
wb=openpyxl.Workbook()
ws=wb.active
ws.append(head)

rec=re.findall("""N-C=.*?(\d+) .*?Nu=\s*([-+]?\d+)\. Uc=  ([-+]?\d+[\.\d]*).*?抗剪承载力""",txt,re.DOTALL)
print(rec)

for r in range(0,3):
        ws['B'+str(2+r)]=rec[r][0]#N-C
        ws['D'+str(2+r)]=rec[r][1]#Nu
        ws['E'+str(2+r)]=rec[r][2]#Uc
wb.save("test.xlsx")
